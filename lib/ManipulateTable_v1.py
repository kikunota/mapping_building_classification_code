import openpyxl
import time
#Get values to retrieve data
#Set values to insert data
#Excel cells start with 1

class ManipulateTable:
    def convertCellUnits(self, col_ceilingH, constant):
        #constant = 1
        new_col_ceilingH = []
        for index in range(len(col_ceilingH)):
            #print (col_ceilingH[index][0])
            val = col_ceilingH[index][0].value
            if type(val) != type(None):
                cell = [int(val) * constant]
            else:
                cell = [val]
            new_col_ceilingH.append(cell)
        return new_col_ceilingH


    def getColumnIndexByName(self, ws, name, headerRow=0):
        """return list of cells"""
        #columnVal = getColumn(header, name)
        #(ws.columns)
        #print(columnVal)

        header = ws.iter_rows(min_row=1, max_col=len(list(ws.rows)[headerRow]), max_row=1, values_only=True)

        for head in header:
            lst_header = list(head)
        #print (lst_header)
        if name in lst_header:
            ind = lst_header.index(name)
        else:
            raise Exception("There is no " + name + " in the header")
        return ind

    def getColumnIndexByName_noValueIncluded(self, ws, name, headerRow=0):
        """return list of cells"""
        #columnVal = getColumn(header, name)
        #(ws.columns)
        #print(columnVal)

        header = ws.iter_rows(min_row=1, max_col=len(list(ws.rows)[headerRow]), max_row=1, values_only=True)

        for head in header:
            lst_header = list(head)
        #print (lst_header)
        if name in lst_header:
            ind = lst_header.index(name)
        else:
            ind = None
            return ind
        return ind

    def getColumnByIndex(self, ws, name, col):
        header = ws.iter_cols(min_col= col, max_row=len(list(ws.columns)[col]), max_col=col, values_only=True)
        for head in header:
            lst_header = list(head)
        lst_header.pop(0)
        return lst_header

    def getColumn(self, header, name):
        for i in range(len(header)):
            if header[i].value == name:
                return i

    def getHeader(self, ws, row=1):
        #return list of cells
        header = list(ws.rows)[row-1]
        return header

    def checkIfRightCol(self, header, headerValue, col):
        if headerValue == header[col-1].value:
            return True
        else:
            return False
    def coordinateToString(self, ex_column, ex_row):
        #make sure row and column is given in excel 1- format
        if ex_column == 0 or ex_row == 0:
            raise Exception("ex_column and ex_row can't be zero")
        coordinate_col = openpyxl.utils.cell.get_column_letter(ex_column)
        coordinate_row = ex_row
        coordinate_str = coordinate_col+str(coordinate_row)

        return coordinate_str

    def convertGeneratorToList(self, nest_generator):
        nestList = []
        for column in nest_generator:
            lst = list(column)
            nestList.append(lst)
        return nestList

    def printCells(self, cells):
        lst = []
        for i in cells:
            lstb = []
            for j in i:
                lstb.append(j.value)
            lst.append(lstb)
        #print(lst)

    def getColumnByName(self, ws, name, headerRow=0):
        #Get column with 部屋名 title
        #名前でその列の情報を落とすことにより列が混ざっていても大丈夫
        col_index = self.getColumnIndexByName(ws,name)
        col_max = len(list(ws.columns)[col_index])
        start_coord = self.coordinateToString(col_index+1, 2)
        end_coord = self.coordinateToString(col_index+1, col_max+1)
        col_room = ws[start_coord: end_coord]
        lst_col_room = self.convertGeneratorToList(col_room)
        return lst_col_room


    def getColumnValueByName(self, ws, name, headerRow=0):
        #Get column with 部屋名 title
        #名前でその列の情報を落とすことにより列が混ざっていても大丈夫
        col_index = self.getColumnIndexByName(ws,name)
        col_max = len(list(ws.columns)[col_index])
        start_coord = self.coordinateToString(col_index+1, 2)
        end_coord = self.coordinateToString(col_index+1, col_max+1)
        col_room = ws[start_coord: end_coord]
        lst_col_room = self.convertGeneratorToList(col_room)
        simple_col_room = self.simplifyList(lst_col_room)
        refined_col_room = self.listCelltoValue(simple_col_room)
        return refined_col_room

    def insertColumn(self, ws, start, columns):
        #Insert Column list with cell.value

        ex_col, ex_row = self.convertCoordinateFromString(start) #retrieve start cell in number
        for i in range(len(columns)):
            ws.cell(row = i + ex_row, column = ex_col, value = columns[i][0].value)

    def insertColumnByValue(self, ws, start, columns):
        #Insert Column list with int or string

        ex_col, ex_row = self.convertCoordinateFromString(start) #retrieve start cell in number
        for i in range(len(columns)):
            ws.cell(row = i + ex_row, column = ex_col, value = columns[i][0])

    def insertColumnByValueB(self, ws, start, columns):
        #Insert Column list with int or string

        ex_col, ex_row = self.convertCoordinateFromString(start) #retrieve start cell in number
        for i in range(len(columns)):
            ws.cell(row = i + ex_row, column = ex_col, value = str(columns[i]))


    def insertID(self, ws, start, columns):
        #This is insert cells
        #Start should be given in cell object

        ex_col, ex_row = self.convertCoordinateFromString(start) #retrieve start cell in number
        for i in range(len(columns)):
            if columns[i][0].value:
                ws.cell(row = i + ex_row, column = ex_col, value = i)

    def insertIDValue(self, ws, start, columns):
        #Insert Column list with strings

        ex_col, ex_row = self.convertCoordinateFromString(start) #retrieve start cell in number
        for i in range(len(columns)):
            if columns[i]:
                ws.cell(row = i + ex_row, column = ex_col, value = str(columns[i]))
    def insertCellValue(self, ws, start, columns):
        #This is insert cells
        #Start should be given in cell object

        ex_col, ex_row = self.convertCoordinateFromString(start) #retrieve start cell in number
        ws.cell(row = ex_row, column = ex_col, value = str(columns))
    def insertCellValueAsNumber(self, ws, start, columns):
        #This is insert cells
        #Start should be given in cell object

        ex_col, ex_row = self.convertCoordinateFromString(start) #retrieve start cell in number
        ws.cell(row = ex_row, column = ex_col, value = columns)

    def convertCoordinateFromString(self, coordinateString):
        col, ex_row = openpyxl.utils.cell.coordinate_from_string(coordinateString)
        ex_col = openpyxl.utils.cell.column_index_from_string(col)
        return ex_col, ex_row

    def simplifyList(self, col):
        simplify_col = [item[0] for item in col]
        return simplify_col

    def listCelltoValue(self, simplifyList):
        list_value = [item.value for item in simplifyList]
        return list_value

    def moveCell(self, cellPos, cellVec):
        pos_col, pos_row = self.convertCoordinateFromString(cellPos)
        mov_col, mov_row = self.convertCoordinateFromString(cellVec)
        pos_col = pos_col + mov_col - 1
        pos_row = pos_row + mov_row - 1
        newCell = self.coordinateToString(pos_col, pos_row)
        return newCell
